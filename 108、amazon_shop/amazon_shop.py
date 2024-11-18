# 爬取亚马逊商品信息
import time
import re
import os
import requests
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 引入隐式等待
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from random import randint


def get_revirews():
    # https://www.amazon.sg/product-reviews/B0CGJDKLB8
    # 打开reviews页面
    rewiew_url = "https://www.amazon.sg/product-reviews" + "/" + ASIN
    browser.implicitly_wait(3)
    browser.get(rewiew_url)
    time.sleep(2)

    review_count = 0
    # 提前打开excel模板
    wb = load_workbook(f"{template_path}/{review_template}")
    ws = wb["Sheet1"]

    review_page = 10
    for i in range(review_page):
        time.sleep(randint(6, 12))
        # 等待页面加载
        WebDriverWait(browser, 4).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, '//div[@id="cm_cr-review_list"]/div[@data-hook="review"]')
            )
        )

        review_list = browser.find_elements(
            By.XPATH, '//div[@id="cm_cr-review_list"]/div[@data-hook="review"]'
        )

        pattern = r"product-reviews/([A-Z0-9]+)"

        for review in review_list:
            review_count += 1
            # 获取评论内容
            review_link = review.find_element(
                By.XPATH,
                './/a[@data-hook="format-strip"]',
            ).get_attribute("href")
            # 使用 re.search 来查找匹配
            match = re.search(pattern, review_link)
            review_asin = match.group(1)
            # print(review_asin)
            review_title = review.find_element(
                By.XPATH, './/span[@data-hook="review-title"]/span'
            ).text
            try:
                review_body = review.find_element(
                    By.XPATH, './/span[@class="cr-original-review-content"]'
                ).text
            except:
                review_body = "No review or Video review"

            review_rating = review.find_element(
                By.XPATH, './/div[@class="a-row a-spacing-none"]/i/span'
            ).text
            # review_rating = review_rating.replace(" out of 5 stars", "")

            review_date = review.find_element(
                By.XPATH, './/span[@data-hook="review-date"]'
            ).text

            profile_name = review.find_element(
                By.XPATH, './/div[@class="a-profile-content"]/span'
            ).text

            profile_link = review.find_element(
                By.XPATH, './/div[@class="a-profile-avatar"]/img'
            ).get_attribute("src")

            # 存入Excel表格
            review_info = [
                review_count,
                review_asin,
                review_title,
                review_body,
                review_rating,
                review_date,
                profile_name,
                profile_link,
            ]

            ws.append(review_info)
            wb.save(f"{file_path}/{ASIN}/{ASIN}_评论信息.xlsx")
            # print(review_info)
            print(f"已存储...{ASIN}...{review_count}条评论")
        # 点击下一页评论，循环获取所有评论
        try:
            browser.find_element(By.XPATH, '//li[@class="a-last"]/a').click()
        except:
            print("No more reviews")
            return


# 获取产品信息
def get_info():
    # 打开商品页面
    target_url = (
        "https://www.amazon.sg/Intel-i7-13700K-Processor-Integrated-Graphics/dp/" + ASIN
    )
    browser.implicitly_wait(4)
    browser.get(target_url)
    time.sleep(2)
    # 设置邮箱
    ## 判断邮编是否一致
    ## //*[@id="glow-ingress-line2"]
    current_zip_code = browser.find_element(By.ID, "glow-ingress-line2").text
    if zip_code not in current_zip_code:

        # //*[@id="glow-ingress-block"]
        browser.find_element(By.ID, "glow-ingress-block").click()

        try:
            browser.find_element(By.ID, "GLUXChangePostalCodeLink").click()
        except:
            pass

        time.sleep(1)
        browser.find_element(By.ID, "GLUXZipUpdateInput").clear()
        browser.find_element(By.ID, "GLUXZipUpdateInput").send_keys(zip_code)
        # //*[@id="GLUXZipUpdate"]/span/input
        browser.find_element(By.ID, "GLUXZipUpdate").click()
        time.sleep(1)
        browser.refresh()
        time.sleep(2)
    print(f"已设置邮编为:{zip_code}")
    # 开始获取图片
    # 悬停至所有图片，让大图加载
    # //*[@id="altImages"]
    alt_iamges_el = browser.find_elements(
        By.XPATH, '//*[@id="altImages"]/ul/li[@data-csa-c-type="uxElement"]//img'
    )
    for img in alt_iamges_el:
        time.sleep(0.3)
        ActionChains(browser).move_to_element(img).perform()

    # # //*[@id="main-image-container"]/ul
    img_list = browser.find_elements(
        By.XPATH,
        '//ul[@class="a-unordered-list a-nostyle a-horizontal list maintain-height"]/li[@data-csa-c-posy]//img',
    )

    for num, img in enumerate(img_list):
        num += 1
        img_link = img.get_attribute("src")
        # https://m.media-amazon.com/images/I/61m0zH-NiTL._AC_SY450_.jpg
        # https://m.media-amazon.com/images/I/61m0zH-NiTL._AC_1500_.jpg
        img_link = re.sub(r"_S.*?.jpg", "_1500_.jpg", img_link)
        # print(img_link)
        image_content = requests.get(img_link).content
        if not os.path.exists(f"{file_path}/{ASIN}"):
            os.mkdir(f"{file_path}/{ASIN}")
        if not os.path.exists(f"{file_path}/{ASIN}/images"):
            os.mkdir(f"{file_path}/{ASIN}/images")

        with open(f"{file_path}/{ASIN}/images/{ASIN}_图片_{num}.jpg", "wb") as f:
            f.write(image_content)
            print(f"已存储{ASIN}....第{num}张图片")

    # 获取产品标题
    # //*[@id="title"]
    product_title = browser.find_element(By.ID, "productTitle").text
    print(f"产品标题为:{product_title}")

    # # 获取价格标题
    try:
        product_whole = browser.find_element(
            By.XPATH,
            '//div[@id="corePriceDisplay_desktop_feature_div"]//span[@class="a-price-whole"]',
        ).text
        product_fraction = browser.find_element(
            By.XPATH,
            '//div[@id="corePriceDisplay_desktop_feature_div"]//span[@class="a-price-fraction"]',
        ).text
        product_price = product_whole + "." + product_fraction
        print(f"产品价格为:{product_price}")
    except:
        product_price = browser.find_element(
            By.XPATH, '//div[@id="corePrice_feature_div"]/span'
        ).text.replace("$", "")
        print(f"产品价格为:{product_price}")

    # 获取优惠券信息
    try:
        voucher_ele = browser.find_element(
            By.XPATH,
            '//div[@id="promoPriceBlockMessage_feature_div"]//span/label',
        ).text
        product_voucher = re.findall("Apply(.*?) voucher", voucher_ele)[0].strip()
        # print(f"优惠券信息为:{voucher_ele}")
        print(f"已获取....{product_voucher}....的优惠券")
    except:
        product_voucher = ""

    # 获取五点描述
    try:
        browser.find_element(
            By.XPATH, '//*[@id="pov2FeatureBulletsExpanderHeading"]/a/span'
        ).click()
    except:
        pass
    # id="featurebullets_feature_div"
    bullet_lists = browser.find_elements(
        By.XPATH,
        '//div[@id="featurebullets_feature_div"]/div/ul/li/span[@class="a-list-item"]',
    )

    bullet_points = []

    for bullet in bullet_lists:
        if bullet.text == "":
            continue
        bullet_points.append(bullet.text)
        # print(bullet.text)

    # 将产品信息存入Exccel表格
    product_info = [
        ASIN,
        product_price,
        product_title,
        product_voucher,
        bullet_points[0],
        bullet_points[1],
        bullet_points[2],
        bullet_points[3],
        bullet_points[4],
    ]

    wb = load_workbook(f"{template_path}/{product_template}")
    ws = wb["Sheet1"]
    ws.append(product_info)
    wb.save(f"{file_path}/{ASIN}/{ASIN}_产品信息.xlsx")
    print(f"已存储{ASIN}....产品信息")


if __name__ == "__main__":
    # 初始化浏览器
    service = Service("D:\Code\chromedriver-win64\chromedriver.exe")
    opt = Options()
    opt.page_load_strategy = "eager"
    opt.debugger_address = "127.0.0.1:8888"
    browser = webdriver.Chrome(service=service, options=opt)
    # 设置参数
    ASIN = "B0CGJDKLB8"
    zip_code = "018926"  # 018926 018906
    file_path = r"D:\Code\DarkCatPython\WebCrawler\amazon_shop"
    template_path = r"D:\Code\DarkCatPython\WebCrawler\amazon_shop\template"
    product_template = "AmazonInfo.xlsx"
    review_template = "AmazonComment.xlsx"

    # get_info()
    get_revirews()
