# 获取淘宝信息并存取
import time
import os
from lxml import etree
from random import uniform
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 引入隐式等待
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook


# 打开搜索结果页
def get_page_info(search_page):
    url = "https://www.taobao.com/"
    browser.get(url)
    browser.implicitly_wait(3)
    time.sleep(3)

    # 复制excle模板
    wb = load_workbook(f"{file_path}/{product_template}")
    wb.save(f"{file_path}/{search_keyword}_商品信息1.xlsx")

    # 输入搜索关键字
    browser.find_element(By.ID, "q").send_keys(search_keyword, Keys.ENTER)

    # 循环翻页
    for i in range(search_page):

        # 等待页面全部加载
        time.sleep(3)
        WebDriverWait(browser, 6).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[3]/div')
            )
        )

        # 获取页面滚动高度
        scroll_height = browser.execute_script("return document.body.scrollHeight")

        # 调整滚动速度
        scroll_step = scroll_height / (2 * 10)  # 每10毫秒滚动的距离
        current_scroll = 0

        # 模拟滚动
        while current_scroll < scroll_height - 1000:
            browser.execute_script(f"window.scrollTo(0, {current_scroll})")
            current_scroll += scroll_step
            time.sleep(0.01)  # 暂停10毫秒

        # 获取页面源代码
        page_text = browser.page_source
        get_product_info(page_text, i + 1)
        # print(page_text)

        time.sleep(uniform(3, 6))
        # 获取下一页
        # //*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/button[2]/span
        browser.find_element(
            By.XPATH,
            '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/button[2]/span',
        ).click()


# 提取产品信息，并保存
def get_product_info(page_text, page_num):
    # 打开excel表格
    wb = load_workbook(f"{file_path}/{search_keyword}_商品信息1.xlsx")
    ws = wb["Sheet1"]

    html = etree.HTML(page_text)
    # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[1]
    item_list = html.xpath('//*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a')
    for item in item_list:
        # 获取产品排名
        index_num = item.xpath(".//@data-spm")[0]
        # 获取商品的标题
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[3]/div/div[1]/div[2]/div/span/text()[1]
        title_list = item.xpath(".//div/div[1]/div[2]/div/span/text()")
        if len(title_list) > 1:
            protect_title = title_list[0] + title_list[1]
        else:
            protect_title = title_list[0] if title_list else "默认标题"

        # 获取产品价格
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[1]/div/div[1]/div[4]/div[1]
        product_price = item.xpath(".//div/div[1]/div[4]/div[1]/span/text()")
        if len(product_price) > 1:
            product_price = product_price[0] + product_price[1]
        else:
            product_price = product_price[0] if product_price else "默认价格"

        # 获取产品付款人数
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[1]/div/div[1]/div[4]/span[2]
        product_deal_count = item.xpath(".//div/div[1]/div[4]/span/text()")
        if len(product_deal_count) > 1:
            product_deal_count = product_deal_count[0] + product_deal_count[1]
        else:
            product_deal_count = (
                product_deal_count[0] if product_deal_count else "默认付款人数"
            )
        # 获取产品店铺名称
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[2]/div/div[3]/div[1]/a/span/span
        product_shop_name = item.xpath(".//div/div[3]/div[1]/a/span/span/text()")[0]
        # 获取地点信息
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[41]/div/div[1]/div[4]/div[2]/span
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[41]/div/div[1]/div[4]/div[3]/span
        product_location = item.xpath(".//div/div[1]/div[4]/div[2]/span/text()")
        if len(product_location) > 1:
            product_location = product_location[0] + product_location[1]
        else:
            product_location = product_location[0] if product_location else "默认地点"
        # 获取产品链接
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[42]
        url = item.xpath(".//@href")[0]
        product_url = "https:" + url

        # 缩略图url
        # //*[@id="search-content-leftWrap"]/div[3]/div[3]/div/a[41]/div/div[1]/div[1]/img
        product_img_url = item.xpath(".//div/div[1]/div[1]/img/@src")[0]
        # 汇总列表中
        product_info = [
            page_num,
            index_num,
            protect_title,
            product_price,
            product_deal_count,
            product_shop_name,
            product_location,
            product_url,
            product_img_url,
        ]
        # 存入excle表格
        ws.append(product_info)
        print(f"已存入...第{page_num}页...第{index_num}个...商品信息")
    # 保存文件
    wb.save(f"{file_path}/{search_keyword}_商品信息1.xlsx")
    # print(product_info)


if __name__ == "__main__":
    # 初始化浏览器
    service = Service("D:\Code\chromedriver-win64\chromedriver.exe")
    opt = Options()
    opt.debugger_address = "127.0.0.1:8888"
    browser = webdriver.Chrome(service=service, options=opt)
    # 配置参数，
    search_keyword = "滑板"
    search_page = 2
    file_path = "D:/Code/DarkCatPython/WebCrawler/taobao_shop"
    product_template = "淘宝产品信息输出表格模板.xlsx"
    # 获取商品信息
    get_page_info(search_page)
