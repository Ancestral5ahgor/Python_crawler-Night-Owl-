# 抓取东方财富网的股票数据
import requests
import os
import pprint
import json
from openpyxl import load_workbook


# 读取股票数据
def get_data(page):
    url = f"https://41.push2.eastmoney.com/api/qt/clist/get?cb=jQuery112408943528881268243_1729148898482&pn={page}&pz=20&po=1&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&dect=1&wbp2u=|0|0|0|web&fid=f3&fs=b:BK0707+f:!50&fields=f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f22,f11,f62,f128,f136,f115,f152,f45&_=1729148898483"
    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    print(response.status_code)
    # 将字符串解析为 Python 字典
    data = response.text.split("(", 1)[1].rstrip(");")
    # print(data)
    json_data = json.loads(data)
    stock_data = parse_data(json_data)
    save_data(stock_data)


# 解析股票数据
def parse_data(data):
    stock_data = []
    stock_list = data["data"]["diff"]
    for stock in stock_list:
        stock_symbol = stock["f12"]
        stock_name = stock["f14"]
        stock_price = stock["f2"]
        stock_percent = stock["f3"]
        stock_updown = stock["f4"]
        stock_open = stock["f17"]
        stock_yeste_close = stock["f18"]
        stock_high = stock["f15"]
        stock_low = stock["f16"]
        stock_volume = f"{stock['f5'] / 10000:.2f}"

        stock_data.append(
            [
                stock_symbol,
                stock_name,
                stock_price,
                stock_percent,
                stock_updown,
                stock_open,
                stock_yeste_close,
                stock_high,
                stock_low,
                stock_volume,
            ]
        )

    return stock_data


# 保存股票数据
def save_data(stock_data):
    wb = load_workbook(f"{file_path}/股票数据.xlsx")
    ws = wb["Sheet1"]
    for stock in stock_data:
        ws.append(stock)
        print("写入成功")
    wb.save(f"{file_path}/股票数据.xlsx")


if __name__ == "__main__":
    file_path = r"D:\Code\DarkCatPython\WebCrawler\eastmoney"
    template_name = "stockinfo.xlsx"

    wb = load_workbook(os.path.join(file_path, template_name))
    wb.save(f"{file_path}/股票数据.xlsx")

    page = 2
    get_data(page)
